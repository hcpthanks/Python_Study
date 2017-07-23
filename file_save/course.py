import csv
import openpyxl

courses = [
    (1, 'python入门', 'hcpthanks', 499),
    (2, 'python入门', 'hcpthanks', 499),
    (3, 'python入门', 'hcpthanks', 499),
    (4, 'python入门', 'hcpthanks', 499),
    (5, 'python入门', 'hcpthanks', 499),
]

# with open('course3.txt', 'w', encoding='utf8') as f:
#     for course in courses:
#         row = ','.join([str(x) for x in course])
#         f.write(f'{row}\n')


# header = ['序号', '课程', '讲师', '定价']
# with open('course2.csv', 'w', encoding='utf8', newline='') as f:
#     write = csv.writer(f)
#     write.writerow(header)
#     write.writerows(courses)

#
# with open('course2.csv','r', encoding='utf8')as f:
#     reader = csv.reader(f)
#     for row in reader:
#         print(row)

# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = '课程'
# ws['E7'] = 'python 爬虫'
# ws.cell(row=1, column=1, value='python')
# wb.save('test.xlsx')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '课程'
ws['A1'] = 'ID'
ws['B1'] = '课程标题'
ws['C1'] = '讲师'
ws['D1'] = '定价'

for i, course in emumerate(courses):
    for j, c in enumerate(course):
        ws.cell(row=i+2, column=j+1, value=c)
wb.save('course1.xlsx')
